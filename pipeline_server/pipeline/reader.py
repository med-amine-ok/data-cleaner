from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import chardet
import pandas as pd
from odf import teletype
from odf.opendocument import load as load_ods
from odf.table import Table, TableCell, TableRow
from openpyxl import load_workbook


class FileIngester:
    """
    Read uploaded spreadsheet or data files into a Pandas DataFrame.

    This class supports CSV, JSON, XLSX, and ODS files. It normalizes the
    input into a DataFrame and forces all values to string form initially so
    later cleaning steps can safely handle messy data.

    Methods:
        read(file_path): Load the file at the given path and return a DataFrame.
    """

    MAX_INGEST_ROWS: int = 100_000
    MAX_REPEAT_EXPANSION: int = 5_000
    MAX_CONSECUTIVE_EMPTY_ROWS: int = 2_000

    def read(self, file_path: str) -> pd.DataFrame:
        """
        Read a file from disk and convert it into a DataFrame.

        Args:
            file_path: Path to the uploaded file on disk.

        Returns:
            A Pandas DataFrame containing the file contents with string columns.

        Raises:
            ValueError: If the file extension is not supported or the file cannot
                be read successfully.
        """
        path = Path(file_path)
        suffix = path.suffix.lower()

        if suffix == ".csv":
            return self._read_csv(path)

        if suffix == ".json":
            return self._read_json(path)

        if suffix == ".xlsx":
            return self._read_excel(path)

        if suffix == ".ods":
            return self._read_ods(path)

        raise ValueError(
            f"Unsupported file format: {suffix}. Supported formats are .csv, .json, .xlsx, and .ods."
        )

    def _read_csv(self, path: Path) -> pd.DataFrame:
        """
        Read a CSV file with encoding detection and string-safe loading.

        Args:
            path: Path to the CSV file.

        Returns:
            A DataFrame with all values loaded as strings.
        """
        raw_bytes = path.read_bytes()
        detected_encoding = self._detect_encoding(raw_bytes)

        candidate_encodings: list[str] = ["utf-8-sig"]
        if detected_encoding and detected_encoding.lower() not in {"utf-8-sig"}:
            candidate_encodings.append(detected_encoding)
        candidate_encodings.extend(["cp1252", "latin-1"])

        tried: set[str] = set()
        best_frame: pd.DataFrame | None = None
        best_score = -1

        for encoding in candidate_encodings:
            normalized_encoding = encoding.lower()
            if normalized_encoding in tried:
                continue
            tried.add(normalized_encoding)

            try:
                frame = pd.read_csv(
                    path,
                    encoding=encoding,
                    dtype=str,
                    keep_default_na=False,
                    na_values=[],
                )
            except UnicodeDecodeError:
                continue

            frame = self._normalize_dataframe(frame)
            score = self._header_quality_score(frame)

            if score > best_score:
                best_score = score
                best_frame = frame

            if score >= 1000:
                break

        if best_frame is None:
            raise ValueError(f"Unable to decode CSV file: {path}")

        return best_frame

    def _read_json(self, path: Path) -> pd.DataFrame:
        """
        Read a JSON file and convert it into a DataFrame.

        Args:
            path: Path to the JSON file.

        Returns:
            A DataFrame with all values loaded as strings.
        """
        raw_text = path.read_text(encoding=self._guess_text_encoding(path))
        parsed: Any = json.loads(raw_text)

        dataframe = pd.DataFrame(parsed)
        return self._normalize_dataframe(dataframe)

    def _read_excel(self, path: Path) -> pd.DataFrame:
        """
        Read an Excel or ODS file and convert it into a DataFrame.

        Args:
            path: Path to the spreadsheet file.

        Returns:
            A DataFrame with all values loaded as strings.
        """
        try:
            dataframe = pd.read_excel(path, dtype=str)
            return self._normalize_dataframe(dataframe)
        except Exception:
            return self._read_xlsx_fallback(path)

    def _read_xlsx_fallback(self, path: Path) -> pd.DataFrame:
        """
        Read XLSX cells using openpyxl when pandas parsing fails.

        Args:
            path: Path to the XLSX file.

        Returns:
            A DataFrame with string-safe values.
        """
        workbook = load_workbook(filename=str(path), data_only=True, read_only=True)
        worksheet = workbook.active

        rows_data: list[list[str]] = []
        consecutive_empty_rows = 0
        for row in worksheet.iter_rows(values_only=True):
            row_values = ["" if value is None else str(value).strip() for value in row]

            if not any(row_values):
                consecutive_empty_rows += 1
                if rows_data and consecutive_empty_rows >= self.MAX_CONSECUTIVE_EMPTY_ROWS:
                    break
                continue

            consecutive_empty_rows = 0
            rows_data.append(row_values)
            if len(rows_data) >= self.MAX_INGEST_ROWS:
                break

        workbook.close()

        if not rows_data:
            return pd.DataFrame()

        max_columns = max((len(row_values) for row_values in rows_data), default=0)
        padded_rows = [row_values + [""] * (max_columns - len(row_values)) for row_values in rows_data]

        header = [value.strip() for value in padded_rows[0]]
        data_rows = padded_rows[1:] if len(padded_rows) > 1 else []
        dataframe = pd.DataFrame(data_rows, columns=header)
        return self._normalize_dataframe(dataframe)

    def _read_ods(self, path: Path) -> pd.DataFrame:
        """
        Read an ODS file as plain text values.

        This avoids hard failures from malformed date cells (for example,
        years out of range) that can occur in strict date parsing paths.

        Args:
            path: Path to the ODS file.

        Returns:
            A DataFrame with all values loaded as strings.
        """
        document = load_ods(str(path))
        tables = document.spreadsheet.getElementsByType(Table)
        if not tables:
            return pd.DataFrame()

        first_table = tables[0]
        rows_data: list[list[str]] = []
        consecutive_empty_rows = 0

        for row in first_table.getElementsByType(TableRow):
            row_repeat = self._safe_repeat(row.getAttribute("numberrowsrepeated"))
            base_row: list[str] = []

            for cell in row.getElementsByType(TableCell):
                col_repeat = self._safe_repeat(cell.getAttribute("numbercolumnsrepeated"))
                cell_text = self._extract_ods_cell_text(cell)
                base_row.extend([cell_text] * col_repeat)

            normalized_row = self._trim_trailing_empty(base_row)
            row_is_empty = not any(normalized_row)

            if row_is_empty:
                if not rows_data:
                    continue
                consecutive_empty_rows += row_repeat
                if consecutive_empty_rows >= self.MAX_CONSECUTIVE_EMPTY_ROWS:
                    break
                continue

            consecutive_empty_rows = 0
            for _ in range(min(row_repeat, self.MAX_REPEAT_EXPANSION)):
                rows_data.append(normalized_row[:])
                if len(rows_data) >= self.MAX_INGEST_ROWS:
                    break

            if len(rows_data) >= self.MAX_INGEST_ROWS:
                break

        if not rows_data:
            return pd.DataFrame()

        max_columns = max((len(row_values) for row_values in rows_data), default=0)
        padded_rows = [row_values + [""] * (max_columns - len(row_values)) for row_values in rows_data]

        header = [value.strip() for value in padded_rows[0]]
        data_rows = padded_rows[1:] if len(padded_rows) > 1 else []

        dataframe = pd.DataFrame(data_rows, columns=header)
        return self._normalize_dataframe(dataframe)

    def _extract_ods_cell_text(self, cell: TableCell) -> str:
        """
        Extract a safe string value from an ODS cell.

        Args:
            cell: ODS table cell.

        Returns:
            Extracted cell text as a string.
        """
        text_value = teletype.extractText(cell).strip()
        if text_value:
            return text_value

        for attr_name in ("stringvalue", "datevalue", "timevalue", "value", "booleanvalue"):
            try:
                attr_value = cell.getAttribute(attr_name)
            except Exception:
                attr_value = None
            if attr_value not in {None, ""}:
                return str(attr_value).strip()

        return ""

    def _safe_repeat(self, value: Any) -> int:
        """
        Convert an ODS repeat attribute into a usable positive integer.

        Args:
            value: Raw repeat value from ODS attributes.

        Returns:
            Repeat count (minimum 1).
        """
        try:
            repeat_count = int(value)
        except (TypeError, ValueError):
            return 1
        if repeat_count <= 0:
            return 1
        return min(repeat_count, self.MAX_REPEAT_EXPANSION)

    def _trim_trailing_empty(self, values: list[str]) -> list[str]:
        """
        Remove trailing empty cells from a row to reduce memory growth.

        Args:
            values: Raw row values.

        Returns:
            Row values without trailing empty strings.
        """
        trimmed = values[:]
        while trimmed and trimmed[-1] == "":
            trimmed.pop()
        return trimmed

    def _normalize_dataframe(self, dataframe: pd.DataFrame) -> pd.DataFrame:
        """
        Normalize a DataFrame so all columns and values are string-based.

        Args:
            dataframe: Input DataFrame.

        Returns:
            Cleaned DataFrame with stripped column names and string values.
        """
        normalized = dataframe.copy()

        normalized.columns = [str(column).strip() for column in normalized.columns]

        for column in normalized.columns:
            normalized[column] = normalized[column].astype(str).replace({"nan": "", "None": ""}).str.strip()

        return normalized

    def _detect_encoding(self, raw_bytes: bytes) -> str:
        """
        Detect the likely text encoding of raw file bytes.

        Args:
            raw_bytes: Raw bytes from the uploaded file.

        Returns:
            A best-effort encoding name, defaulting to UTF-8 if detection fails.
        """
        detection = chardet.detect(raw_bytes)
        encoding = detection.get("encoding")
        return encoding or "utf-8"

    def _guess_text_encoding(self, path: Path) -> str:
        """
        Guess the encoding of a text-based file using a small sample.

        Args:
            path: Path to the text file.

        Returns:
            A best-effort encoding name.
        """
        raw_bytes = path.read_bytes()
        return self._detect_encoding(raw_bytes)

    def _header_quality_score(self, dataframe: pd.DataFrame) -> int:
        """
        Score dataframe headers to pick the cleanest CSV decoding.

        Higher is better. Headers containing replacement chars are heavily penalized.
        """
        if dataframe.empty and len(dataframe.columns) == 0:
            return -10_000

        score = 0
        for column in dataframe.columns:
            text = str(column)
            score += 100
            score -= text.count("\ufffd") * 500
            score -= text.count("?") * 20
        return score